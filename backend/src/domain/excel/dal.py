import io
import uuid

from openpyxl import load_workbook, Workbook
from pydantic import ValidationError

from src.domain.excel.dto import ExcelImportResultDTO
from src.domain.fish.model import Fish
from src.domain.fish.service import get_fishes_list
from src.domain.group.model import Group


async def read_group_from_excel(file: io.BytesIO):
    workbook = load_workbook(filename=file)
    sheet = workbook.active

    group = Group(breed='Лосось', sex='М', id='', father_group=None, mother_group=None)
    correct_fish_count = 0
    error_fish_count = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        try:
            fish = Fish(
                id=row[0] if row[0] is not None else str(uuid.uuid4()),
                weight=row[1],
                length=row[2],
                height=row[3],
                thickness=row[4],
                eggs_weight=row[5],
                egg_weight=row[6],
            )
            if not group.id:
                group.id = row[7]
            group.fishes.append(fish)
            correct_fish_count += 1
        except ValidationError as e:
            print(e)
            error_fish_count += 1
    await group.save()
    return ExcelImportResultDTO(
        correct_fish_count=correct_fish_count,
        error_fish_count=error_fish_count,
    )


async def write_group_to_excel(group_id: str):
    group = await Group.get(group_id)
    if group is None:
        return None
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = group.breed
    sheet.append(
        [
            'Индентификатор',
            'Вес',
            'Длина',
            'Высота',
            'Толщина',
            'Вес икры',
            'Вес икринки',

            'К упитанности',
            'И толщины',
            'И высоты',
            'Доля икры',
            'Рабочая плодовитость',
            'Относительная плодовитость',
            'Репродуктивный индекс',

            'Шанс хорошего потомства',
            'Группа'
        ]
    )
    fishes = await get_fishes_list(group_id)
    for fish in fishes:
        sheet.append([
            fish.id,
            fish.weight,
            fish.length,
            fish.height,
            fish.thickness,
            fish.eggs_weight,
            fish.egg_weight,

            fish.k_upit,
            fish.i_tolsh,
            fish.i_visots,
            fish.dolya_icry,
            fish.work_plodovitost,
            fish.otnosit_plodovitost,
            fish.index_reproduction,

            fish.predict_proba,
            group.id,
        ])
    file = io.BytesIO()
    workbook.save(file)
    return file


async def write_generations_to_excel():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Родословная индивидуальная'
    sheet.append(
        [
            'ID животного',
            'ID самки',
            'ID самца'
        ]
    )
    groups = await Group.find_all().to_list()
    fishes = [fish for group in groups for fish in group.fishes]
    for fish in fishes:
        sheet.append([
            fish.id,
            fish.mother_id,
            fish.father_id,
        ])
    sheet = workbook.create_sheet(title='Родословная групповая')
    sheet.append(
        [
            'ID группы',
            'ID группы самок',
            'ID группы самцов',
        ]
    )
    for group in groups:
        sheet.append([
            group.id,
            group.mother_group.ref.id,
            group.father_group.ref.id,
        ])
    file = io.BytesIO()
    workbook.save(file)
    return file
